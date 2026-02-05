"""
Import Service Module
Contains: Bulk upload from Excel/CSV with ZIP photo matching
"""
import os
import re
import json
import zipfile
from io import BytesIO
from datetime import datetime, timedelta
from typing import Dict, Any, List, Optional, Tuple

from django.shortcuts import get_object_or_404
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile

from ..models import IDCardTable, IDCard
from .base import BaseService, ServiceResult
from .image_service import ImageService


class ImportService(BaseService):
    """
    Service for bulk importing ID Cards from Excel/CSV files.
    
    Features:
    - XLSX support (openpyxl)
    - XLS support (xlrd)
    - CSV support
    - Fuzzy header matching
    - ZIP photo matching by reference column
    - Multiple image fields support
    """
    
    @classmethod
    def bulk_upload(
        cls,
        table_id: int,
        data_file,
        zip_files: Dict[str, Any] = None,
        zip_field_names: List[str] = None
    ) -> ServiceResult:
        """
        Bulk upload ID Cards from Excel/CSV with optional photo ZIPs.
        
        Args:
            table_id: ID of the target table
            data_file: Uploaded Excel/CSV file
            zip_files: Dict mapping field names to uploaded ZIP files
            zip_field_names: List of field names that have ZIP files
        
        Returns:
            ServiceResult with upload statistics
        """
        try:
            table = get_object_or_404(IDCardTable, id=table_id)
            client = table.group.client
            client_image_folder = ImageService.get_client_image_folder(client)
            
            file_name = data_file.name.lower()
            
            # Get field configurations
            text_fields = [f['name'] for f in table.fields if not cls.is_image_field(f)]
            image_fields = cls.get_image_field_names(table.fields)
            
            # Extract photos from ZIP files
            zip_photos_by_field = cls._extract_zip_photos(
                zip_files or {}, 
                zip_field_names or [],
                image_fields
            )
            
            # Parse data file
            if file_name.endswith(('.xlsx', '.xls')):
                parse_result = cls._parse_excel(data_file, text_fields, image_fields)
            elif file_name.endswith('.csv'):
                parse_result = cls._parse_csv(data_file, text_fields, image_fields)
            else:
                return ServiceResult(
                    success=False,
                    message='Invalid file format! Please upload .xlsx, .xls, or .csv file.'
                )
            
            if not parse_result['success']:
                return ServiceResult(success=False, message=parse_result['message'])
            
            # Create cards
            cards_created = 0
            total_photos_matched = 0
            errors = []
            
            for row_num, row_data in enumerate(parse_result['rows'], start=2):
                try:
                    field_data = row_data['field_data']
                    image_refs = row_data.get('image_refs', {})
                    
                    # Match and save photos
                    photos_matched = cls._match_and_save_photos(
                        field_data,
                        image_refs,
                        image_fields,
                        zip_photos_by_field,
                        client_image_folder,
                        cards_created
                    )
                    total_photos_matched += photos_matched
                    
                    # Create the card
                    IDCard.objects.create(
                        table=table,
                        field_data=field_data,
                        status='pending'
                    )
                    cards_created += 1
                    
                except Exception as e:
                    errors.append(f'Row {row_num}: {str(e)}')
            
            # Build result message
            photo_msg = f" with {total_photos_matched} photos matched" if total_photos_matched > 0 else ""
            result = {
                'cards_created': cards_created,
                'photos_matched': total_photos_matched,
                'matched_fields': parse_result.get('matched_fields', []),
            }
            
            if errors:
                result['errors'] = errors[:10]
                result['error_count'] = len(errors)
            
            return ServiceResult(
                success=True,
                message=f'Successfully created {cards_created} ID cards{photo_msg}!',
                data=result
            )
            
        except ImportError as e:
            return ServiceResult(
                success=False,
                message=f'Required library not installed: {str(e)}'
            )
        except Exception as e:
            return ServiceResult(success=False, message=str(e))
    
    @classmethod
    def _extract_zip_photos(
        cls,
        zip_files: Dict[str, Any],
        zip_field_names: List[str],
        image_fields: List[str]
    ) -> Dict[str, Dict[str, Any]]:
        """Extract photos from ZIP files into memory"""
        zip_photos_by_field = {}
        
        for field_name in zip_field_names:
            zip_key = f'photos_zip_{field_name}'
            if zip_key not in zip_files:
                continue
            
            zip_file = zip_files[zip_key]
            zip_photos_by_field[field_name] = {}
            
            try:
                zip_content = zip_file.read()
                with zipfile.ZipFile(BytesIO(zip_content), 'r') as zf:
                    for zip_info in zf.infolist():
                        if zip_info.is_dir():
                            continue
                        
                        base_name = os.path.basename(zip_info.filename)
                        name_without_ext = os.path.splitext(base_name)[0]
                        ext = os.path.splitext(base_name)[1].lower()
                        
                        if ext in cls.VALID_IMAGE_EXTENSIONS:
                            try:
                                image_bytes = zf.read(zip_info.filename)
                                is_valid, _ = ImageService.validate_image_bytes(image_bytes)
                                
                                if is_valid:
                                    # Store with uppercase key for case-insensitive matching
                                    zip_photos_by_field[field_name][name_without_ext.upper()] = {
                                        'bytes': image_bytes,
                                        'ext': ext
                                    }
                            except Exception:
                                continue
            except Exception:
                pass
        
        return zip_photos_by_field
    
    @classmethod
    def _parse_excel(
        cls,
        data_file,
        text_fields: List[str],
        image_fields: List[str]
    ) -> Dict[str, Any]:
        """Parse Excel file (XLSX or XLS)"""
        try:
            import openpyxl
            
            file_content = data_file.read()
            
            if len(file_content) < 4:
                return {'success': False, 'message': 'File is too small or empty.'}
            
            # Detect format
            magic_bytes = file_content[:4]
            is_zip = magic_bytes[:2] == b'PK'
            is_old_xls = (magic_bytes[0] == 0xD0 and magic_bytes[1] == 0xCF)
            
            headers = []
            rows_data = []
            
            if is_zip or data_file.name.lower().endswith('.xlsx'):
                # XLSX format
                try:
                    wb = openpyxl.load_workbook(BytesIO(file_content))
                    ws = wb.active
                    
                    for cell in ws[1]:
                        if cell.value:
                            headers.append(str(cell.value).strip())
                    
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        rows_data.append(row)
                except Exception as e:
                    if not is_zip:
                        is_old_xls = True
                    else:
                        return {'success': False, 'message': f'Error reading XLSX: {str(e)}'}
            
            if is_old_xls or (data_file.name.lower().endswith('.xls') and not headers):
                # XLS format
                try:
                    import xlrd
                    wb = xlrd.open_workbook(file_contents=file_content)
                    ws = wb.sheet_by_index(0)
                    
                    headers = []
                    for col_idx in range(ws.ncols):
                        cell_value = ws.cell_value(0, col_idx)
                        if cell_value:
                            headers.append(str(cell_value).strip())
                    
                    rows_data = []
                    for row_idx in range(1, ws.nrows):
                        row = []
                        for col_idx in range(ws.ncols):
                            row.append(ws.cell_value(row_idx, col_idx))
                        rows_data.append(tuple(row))
                except ImportError:
                    return {
                        'success': False,
                        'message': 'xlrd library not installed for .xls support'
                    }
                except Exception as e:
                    return {'success': False, 'message': f'Error reading XLS: {str(e)}'}
            
            if not headers:
                return {'success': False, 'message': 'Could not read headers from file.'}
            
            return cls._process_parsed_data(headers, rows_data, text_fields, image_fields)
            
        except Exception as e:
            return {'success': False, 'message': str(e)}
    
    @classmethod
    def _parse_csv(
        cls,
        data_file,
        text_fields: List[str],
        image_fields: List[str]
    ) -> Dict[str, Any]:
        """Parse CSV file"""
        try:
            import csv
            from io import StringIO
            
            content = data_file.read().decode('utf-8-sig')
            reader = csv.DictReader(StringIO(content))
            
            headers = reader.fieldnames or []
            rows_data = [(tuple(row.get(h, '') for h in headers)) for row in reader]
            
            return cls._process_parsed_data(headers, rows_data, text_fields, image_fields)
            
        except Exception as e:
            return {'success': False, 'message': str(e)}
    
    @classmethod
    def _process_parsed_data(
        cls,
        headers: List[str],
        rows_data: List[tuple],
        text_fields: List[str],
        image_fields: List[str]
    ) -> Dict[str, Any]:
        """Process parsed data to build field mappings"""
        
        # Map headers to fields
        header_to_field = {}
        image_ref_columns = {}
        available_fields = text_fields.copy()
        matched_field_names = []
        
        for idx, header in enumerate(headers):
            if not header:
                continue
            
            header_upper = header.upper()
            
            # Check if this is an image reference column
            is_image_ref = False
            for img_field in image_fields:
                if cls.normalize_name(header) == cls.normalize_name(img_field):
                    image_ref_columns[img_field] = idx
                    is_image_ref = True
                    break
            
            if is_image_ref:
                continue
            
            # Try to match to text field
            match = cls.find_best_field_match(header.strip(), available_fields)
            if match:
                header_to_field[idx] = match
                available_fields.remove(match)
                matched_field_names.append(match)
        
        if not header_to_field:
            return {
                'success': False,
                'message': f'No matching columns found! Expected: {", ".join(text_fields)}'
            }
        
        # Process rows
        processed_rows = []
        for row in rows_data:
            # Skip empty rows
            if all(cell is None or str(cell).strip() == '' for cell in row):
                continue
            
            field_data = {}
            image_refs = {}
            
            # Extract text field values
            for col_idx, field_name in header_to_field.items():
                if col_idx < len(row):
                    value = cls._convert_cell_value(row[col_idx], field_name)
                    field_data[field_name] = value
                else:
                    field_data[field_name] = ''
            
            # Extract image reference values
            for img_field, col_idx in image_ref_columns.items():
                if col_idx < len(row):
                    value = row[col_idx]
                    if value is not None and str(value).strip():
                        if isinstance(value, float) and value == int(value):
                            image_refs[img_field] = str(int(value))
                        else:
                            image_refs[img_field] = str(value).strip()
            
            processed_rows.append({
                'field_data': field_data,
                'image_refs': image_refs
            })
        
        return {
            'success': True,
            'rows': processed_rows,
            'matched_fields': matched_field_names
        }
    
    @classmethod
    def _convert_cell_value(cls, value: Any, field_name: str) -> str:
        """Convert cell value to string, handling dates and numbers"""
        if value is None:
            return ''
        
        # Handle dates
        if hasattr(value, 'strftime'):
            return value.strftime('%d-%m-%Y')
        
        # Handle Excel date serial numbers
        if isinstance(value, (int, float)):
            is_date_field = any(
                kw in field_name.lower() 
                for kw in ['date', 'dob', 'birth']
            )
            
            if is_date_field and 1 < value < 60000:
                try:
                    excel_epoch = datetime(1899, 12, 30)
                    actual_date = excel_epoch + timedelta(days=int(value))
                    return actual_date.strftime('%d-%m-%Y')
                except Exception:
                    pass
            
            if isinstance(value, float) and value == int(value):
                return str(int(value)).upper()
            return str(value).upper()
        
        return str(value).strip().upper()
    
    @classmethod
    def _match_and_save_photos(
        cls,
        field_data: Dict[str, Any],
        image_refs: Dict[str, str],
        image_fields: List[str],
        zip_photos_by_field: Dict[str, Dict[str, Any]],
        client_image_folder: str,
        batch_counter: int
    ) -> int:
        """Match and save photos from ZIP files"""
        photos_matched = 0
        
        for img_field in image_fields:
            photo_ref = image_refs.get(img_field, '')
            photo_key = photo_ref.upper() if photo_ref else None
            
            field_zip_photos = zip_photos_by_field.get(img_field, {})
            
            if photo_key and photo_key in field_zip_photos:
                try:
                    photo_info = field_zip_photos[photo_key]
                    
                    new_filename = ImageService.generate_filename(
                        batch_counter + 1,
                        photo_info['ext']
                    )
                    
                    file_path = f"{client_image_folder}/{new_filename}"
                    
                    saved_path = default_storage.save(
                        file_path,
                        ContentFile(photo_info['bytes'])
                    )
                    
                    field_data[img_field] = saved_path
                    photos_matched += 1
                except Exception:
                    if photo_ref:
                        field_data[img_field] = f'PENDING:{photo_ref}'
                    else:
                        field_data[img_field] = ''
            else:
                if photo_ref:
                    field_data[img_field] = f'PENDING:{photo_ref}'
                else:
                    field_data[img_field] = ''
        
        return photos_matched
    
    @classmethod
    def reupload_images(
        cls,
        table_id: int,
        zip_file,
        card_ids: List[int] = None
    ) -> ServiceResult:
        """
        Reupload images from ZIP file, matching by stored filename.
        
        Args:
            table_id: Target table ID
            zip_file: Uploaded ZIP file
            card_ids: Optional list of card IDs to process
        
        Returns:
            ServiceResult with reupload statistics
        """
        try:
            table = get_object_or_404(IDCardTable, id=table_id)
            client = table.group.client
            client_image_folder = ImageService.get_client_image_folder(client)
            
            # Get cards
            if card_ids:
                cards = IDCard.objects.filter(table=table, id__in=card_ids).order_by('id')
            else:
                cards = IDCard.objects.filter(table=table).order_by('id')
            
            if not cards.exists():
                return ServiceResult(success=False, message='No cards found!')
            
            # Get image fields
            image_fields = cls.get_image_field_names(table.fields or [])
            if not image_fields:
                return ServiceResult(
                    success=False, 
                    message='No image fields configured!'
                )
            
            # Extract photos from ZIP
            zip_photos = {}
            invalid_images = 0
            
            zip_content = zip_file.read()
            with zipfile.ZipFile(BytesIO(zip_content), 'r') as zf:
                for zip_info in zf.infolist():
                    if zip_info.is_dir():
                        continue
                    
                    base_name = os.path.basename(zip_info.filename)
                    name_without_ext = os.path.splitext(base_name)[0]
                    ext = os.path.splitext(base_name)[1].lower()
                    
                    if ext in cls.VALID_IMAGE_EXTENSIONS:
                        try:
                            image_bytes = zf.read(zip_info.filename)
                            is_valid, _ = ImageService.validate_image_bytes(image_bytes)
                            
                            if is_valid:
                                zip_photos[name_without_ext] = {
                                    'bytes': image_bytes,
                                    'ext': ext,
                                    'original_name': base_name
                                }
                            else:
                                invalid_images += 1
                        except Exception:
                            invalid_images += 1
            
            if not zip_photos:
                return ServiceResult(
                    success=False, 
                    message='No valid images found in ZIP!'
                )
            
            # Process cards
            images_matched = 0
            cards_updated = 0
            batch_counter = 0
            
            for card in cards:
                field_data = card.field_data or {}
                card_updated = False
                
                for img_field in image_fields:
                    current_value = field_data.get(img_field, '')
                    if not current_value or not current_value.strip():
                        continue
                    
                    # Build matching keys
                    matching_keys = cls._get_matching_keys(current_value)
                    
                    # Find match
                    matched_photo = None
                    for key in matching_keys:
                        if key in zip_photos:
                            matched_photo = zip_photos[key]
                            break
                    
                    if matched_photo:
                        batch_counter += 1
                        
                        # Determine if update or new
                        existing_path = None
                        if current_value not in ['NOT_FOUND', 'PENDING'] and '/' in current_value:
                            existing_path = current_value
                        
                        if existing_path:
                            new_filename = ImageService.generate_updated_filename(
                                existing_path, 
                                matched_photo['ext']
                            )
                            # Delete old
                            ImageService.delete_image(existing_path)
                        else:
                            new_filename = ImageService.generate_filename(
                                batch_counter,
                                matched_photo['ext']
                            )
                        
                        file_path = f"{client_image_folder}/{new_filename}"
                        
                        saved_path = default_storage.save(
                            file_path,
                            ContentFile(matched_photo['bytes'])
                        )
                        
                        field_data[img_field] = saved_path
                        card_updated = True
                        images_matched += 1
                
                if card_updated:
                    card.field_data = field_data
                    card.save()
                    cards_updated += 1
            
            if images_matched == 0:
                return ServiceResult(
                    success=False,
                    message=f'No matching images found! ZIP has {len(zip_photos)} images.'
                )
            
            return ServiceResult(
                success=True,
                message=f'Reuploaded {images_matched} images for {cards_updated} cards!',
                data={
                    'images_matched': images_matched,
                    'cards_updated': cards_updated,
                    'invalid_images': invalid_images,
                    'zip_image_count': len(zip_photos)
                }
            )
            
        except Exception as e:
            return ServiceResult(success=False, message=str(e))
    
    @staticmethod
    def _get_matching_keys(current_value: str) -> List[str]:
        """Get list of possible matching keys for an image value"""
        matching_keys = []
        
        if current_value.startswith('PENDING:'):
            pending_ref = current_value[8:]
            ref_no_ext = os.path.splitext(pending_ref)[0] if '.' in pending_ref else pending_ref
            if ref_no_ext:
                matching_keys.append(ref_no_ext)
            if pending_ref:
                matching_keys.append(pending_ref)
        elif '/' in current_value:
            stored_filename = os.path.splitext(os.path.basename(current_value))[0]
            if stored_filename:
                matching_keys.append(stored_filename)
        else:
            ref_name = os.path.splitext(current_value)[0] if '.' in current_value else current_value
            if ref_name:
                matching_keys.append(ref_name)
        
        return matching_keys

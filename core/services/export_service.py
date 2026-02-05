"""
Export Service Module
Contains: DOCX, XLSX, and ZIP export operations for ID Cards
"""
import os
import base64
import re
from io import BytesIO
from datetime import datetime
from typing import Dict, Any, List, Optional

from django.shortcuts import get_object_or_404
from django.core.files.storage import default_storage
from django.http import HttpResponse

from ..models import IDCardTable, IDCard
from .base import BaseService, ServiceResult


class ExportService(BaseService):
    """
    Service for exporting ID Card data in various formats.
    
    Supported formats:
    - DOCX: Word document with images and table layout
    - XLSX: Excel spreadsheet (text fields only)
    - ZIP: Images organized by field name
    """
    
    ENTRIES_PER_PAGE = 7  # Cards per page in DOCX
    
    @classmethod
    def export_xlsx(
        cls, 
        table_id: int, 
        card_ids: List[int]
    ) -> ServiceResult:
        """
        Export selected cards as Excel file.
        
        Returns:
            ServiceResult with 'response' key containing HttpResponse
        """
        try:
            import openpyxl
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            
            table = get_object_or_404(IDCardTable, id=table_id)
            
            if not card_ids:
                return ServiceResult(success=False, message='No cards selected!')
            
            cards = IDCard.objects.filter(table=table, id__in=card_ids).order_by('id')
            if not cards.exists():
                return ServiceResult(success=False, message='No cards found!')
            
            # Get text fields only
            text_fields = cls.get_text_fields(table.fields or [])
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = table.name[:31]
            
            # Styles
            header_font = Font(name='Calibri', size=11, bold=True)
            header_alignment = Alignment(horizontal='center', vertical='center')
            data_font = Font(name='Calibri', size=10)
            data_alignment = Alignment(horizontal='left', vertical='center')
            thin_border = Border(
                left=Side(style='thin', color='CCCCCC'),
                right=Side(style='thin', color='CCCCCC'),
                top=Side(style='thin', color='CCCCCC'),
                bottom=Side(style='thin', color='CCCCCC')
            )
            
            # Track column widths
            column_widths = {}
            
            # Header row
            headers = [f['name'] for f in text_fields]
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = thin_border
                column_widths[col_idx] = len(str(header)) + 2
            
            # Data rows
            for row_idx, card in enumerate(cards, 2):
                field_data = card.field_data or {}
                
                for col_idx, field in enumerate(text_fields, 1):
                    value = field_data.get(field['name'], '')
                    value_str = str(value).upper() if value else ''
                    
                    cell = ws.cell(row=row_idx, column=col_idx, value=value_str)
                    cell.font = data_font
                    cell.alignment = data_alignment
                    cell.border = thin_border
                    
                    # Track max width
                    current_width = min(len(value_str) + 2, 50)
                    column_widths[col_idx] = max(column_widths.get(col_idx, 8), current_width)
            
            # Apply column widths
            for col_idx, width in column_widths.items():
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = max(8, width * 1.1)
            
            ws.row_dimensions[1].height = 25
            ws.freeze_panes = 'A2'
            
            # Save to buffer
            xlsx_buffer = BytesIO()
            wb.save(xlsx_buffer)
            xlsx_buffer.seek(0)
            
            # Generate filename
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"{table.name}_{timestamp}.xlsx"
            
            response = HttpResponse(
                xlsx_buffer.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            response['Content-Length'] = len(xlsx_buffer.getvalue())
            
            return ServiceResult(
                success=True,
                data={'response': response}
            )
            
        except ImportError:
            return ServiceResult(
                success=False,
                message='openpyxl library not installed. Run: pip install openpyxl'
            )
        except Exception as e:
            return ServiceResult(success=False, message=str(e))
    
    @classmethod
    def export_images_zip(
        cls, 
        table_id: int, 
        card_ids: List[int]
    ) -> ServiceResult:
        """
        Export images as separate ZIP files for each image column.
        
        Returns:
            ServiceResult with 'zip_files' list (base64 encoded)
        """
        try:
            import zipfile
            
            table = get_object_or_404(IDCardTable, id=table_id)
            
            if not card_ids:
                return ServiceResult(success=False, message='No cards selected!')
            
            cards = IDCard.objects.filter(table=table, id__in=card_ids).order_by('id')
            if not cards.exists():
                return ServiceResult(success=False, message='No cards found!')
            
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            
            # Get image fields
            image_fields = cls.get_image_field_names(table.fields or [])
            
            if not image_fields:
                return ServiceResult(
                    success=False, 
                    message='No image fields found in this table!'
                )
            
            zip_files = []
            total_images = 0
            
            for img_field in image_fields:
                zip_buffer = BytesIO()
                images_in_field = 0
                
                with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
                    for card in cards:
                        field_data = card.field_data or {}
                        img_path = field_data.get(img_field, '')
                        
                        if img_path and img_path != 'NOT_FOUND' and img_path.strip():
                            try:
                                if default_storage.exists(img_path):
                                    with default_storage.open(img_path, 'rb') as img_file:
                                        img_data = img_file.read()
                                        
                                        if img_data and len(img_data) >= 100:
                                            download_filename = os.path.basename(img_path)
                                            zf.writestr(download_filename, img_data)
                                            images_in_field += 1
                            except Exception:
                                continue
                
                if images_in_field > 0:
                    zip_buffer.seek(0)
                    zip_data = zip_buffer.getvalue()
                    
                    # Clean field name for filename
                    clean_field_name = cls._get_readable_field_name(img_field)
                    clean_table_name = cls.clean_filename_for_export(table.name)
                    
                    zip_filename = f"{clean_table_name}_{clean_field_name}_{timestamp}.zip"
                    zip_base64 = base64.b64encode(zip_data).decode('utf-8')
                    
                    zip_files.append({
                        'field_name': img_field,
                        'filename': zip_filename,
                        'data': zip_base64,
                        'image_count': images_in_field
                    })
                    total_images += images_in_field
            
            if not zip_files:
                return ServiceResult(
                    success=False, 
                    message='No images found for selected cards!'
                )
            
            return ServiceResult(
                success=True,
                data={
                    'zip_files': zip_files,
                    'total_images': total_images,
                    'total_zips': len(zip_files)
                }
            )
            
        except Exception as e:
            return ServiceResult(success=False, message=str(e))
    
    @staticmethod
    def _get_readable_field_name(field_name: str) -> str:
        """Convert field name to readable format for filename"""
        name_upper = field_name.upper().strip()
        mappings = {
            'F PHOTO': 'FATHER_PHOTO',
            'M PHOTO': 'MOTHER_PHOTO',
            'SIGN': 'SIGNATURE',
        }
        return mappings.get(name_upper, name_upper.replace(' ', '_'))
    
    @classmethod
    def export_docx(
        cls, 
        table_id: int, 
        card_ids: List[int],
        doc_format: str = 'docx'
    ) -> ServiceResult:
        """
        Export selected cards as Word document.
        
        This is a complex operation - the full implementation is in the
        original view. This provides a cleaner interface.
        
        Returns:
            ServiceResult with 'response' key containing HttpResponse
        """
        # NOTE: The DOCX export is quite complex (300+ lines of formatting).
        # For a full refactor, consider keeping this in a separate module
        # or maintaining the current implementation for now.
        # 
        # Key points for refactoring:
        # 1. Document setup (page orientation, margins)
        # 2. Header/footer generation
        # 3. Table creation with dynamic columns
        # 4. Image embedding with proper sizing
        # 5. Pagination (7 entries per page)
        
        try:
            from docx import Document
            from docx.shared import Inches, Cm, Pt, RGBColor
            from docx.enum.table import WD_TABLE_ALIGNMENT
            from docx.enum.text import WD_ALIGN_PARAGRAPH
            
            table = get_object_or_404(IDCardTable, id=table_id)
            
            if not card_ids:
                return ServiceResult(success=False, message='No cards selected!')
            
            cards = IDCard.objects.filter(table=table, id__in=card_ids).order_by('id')
            if not cards.exists():
                return ServiceResult(success=False, message='No cards found!')
            
            # For full implementation, call the original view function
            # or implement the full DOCX generation here
            
            return ServiceResult(
                success=False,
                message='DOCX export requires full implementation - use original view'
            )
            
        except ImportError:
            return ServiceResult(
                success=False,
                message='python-docx library not installed. Run: pip install python-docx'
            )
        except Exception as e:
            return ServiceResult(success=False, message=str(e))
